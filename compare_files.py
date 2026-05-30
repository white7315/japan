import os
import openpyxl

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
src_file = "2026년 05월 매출일지.xlsx"
dest_file = "26년5월 매출일지.xlsx"

src_path = os.path.join(desktop, src_file)
dest_path = os.path.join(desktop, dest_file)

wb_src = openpyxl.load_workbook(src_path, data_only=True)
wb_dest = openpyxl.load_workbook(dest_path, data_only=True)

print("--- 누적 합계 (Row 28) 및 일일 합계 비교 ---")
for sheet_name in wb_src.sheetnames:
    if "5월 종합" in sheet_name:
        continue
    sheet_src = wb_src[sheet_name]
    sheet_dest = wb_dest[sheet_name]
    
    # 28행 열별 값 가져오기
    # Col 9 (I): 누적 카드
    # Col 10 (J): 누적 현금
    # Col 11 (K): 누적 총합
    # Col 12 (L): 누적 총합 (일부 시트)
    # Col 15 (O): 누적 마진
    
    diffs = []
    for col_idx in [9, 10, 11, 12, 15]:
        val_src = sheet_src.cell(28, col_idx).value
        val_dest = sheet_dest.cell(28, col_idx).value
        
        # 엑셀과 구글시트 공백 처리 차이 감안
        val_src_num = val_src if isinstance(val_src, (int, float)) else 0
        val_dest_num = val_dest if isinstance(val_dest, (int, float)) else 0
        
        if val_src_num != val_dest_num:
            diffs.append(f"Col {col_idx}({openpyxl.utils.get_column_letter(col_idx)}): 원본={val_src}, 새파일={val_dest}")
            
    if diffs:
        print(f"[{sheet_name}] 누적합계 불일치:")
        for d in diffs:
            print(f"  {d}")

wb_src.close()
wb_dest.close()
