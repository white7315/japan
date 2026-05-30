import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
may_file = os.path.join(desktop, "26년5월 매출일지.xlsx")

wb = openpyxl.load_workbook(may_file, data_only=False)
sheet = wb["5월1일"]

print("--- Sheet Info for 5월1일 ---")
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

print("\n--- Merged Ranges ---")
for r in sorted(sheet.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
    print(f"  {r}")

print("\n--- Non-empty cells/formulas in 5월1일 ---")
for r in range(1, 35):
    row_vals = []
    for c in range(1, 17):
        cell = sheet.cell(r, c)
        if cell.value is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            # check style properties
            font_color = cell.font.color.rgb if cell.font and cell.font.color else None
            font_bold = cell.font.bold if cell.font else None
            font_size = cell.font.size if cell.font else None
            align_horiz = cell.alignment.horizontal if cell.alignment else None
            number_format = cell.number_format
            row_vals.append(f"{col_letter}{r}: '{cell.value}' [bold={font_bold}, size={font_size}, color={font_color}, align={align_horiz}, fmt={number_format}]")
    if row_vals:
        print(f"Row {r}: " + " | ".join(row_vals))

print("\n--- Column Widths in 5월1일 ---")
for col_idx in range(1, 17):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    dim = sheet.column_dimensions[col_letter]
    print(f"Col {col_letter}: width={dim.width}")

print("\n--- Row Heights in 5월1일 ---")
for r in range(1, 32):
    dim = sheet.row_dimensions[r]
    print(f"Row {r}: height={dim.height}")

wb.close()
