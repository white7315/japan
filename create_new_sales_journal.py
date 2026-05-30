import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
src_file = None
for name in os.listdir(desktop):
    if "2026" in name and "05" in name and name.endswith(".xlsx") and name != "26년5월 매출일지.xlsx":
        src_file = name
        break

if not src_file:
    print("원본 엑셀 파일을 찾지 못했습니다.")
    exit()

src_path = os.path.join(desktop, src_file)
dest_path = os.path.join(desktop, "26년5월 매출일지.xlsx")

# 1. 원본 파일에서 시트 구조 분석 (data_only=False로 수식 스캔)
wb_formulas = openpyxl.load_workbook(src_path, data_only=False)
original_sum_rows = {}

def find_sum_row_formulas(sheet):
    # B열에서 "합 계" 텍스트 탐색
    for r in range(6, 40):
        val_b = sheet.cell(r, 2).value
        if val_b and isinstance(val_b, str) and "합" in val_b and "계" in val_b:
            return r
            
    # 주요 수치/합계 열(E, F, G, I, J, K, L, N)에서 =SUM( 수식 탐색
    for r in range(6, 40):
        for c in [5, 6, 7, 9, 10, 11, 12, 14]:
            val = sheet.cell(r, c).value
            if val and isinstance(val, str) and "=SUM(" in val.replace(" ", "").upper():
                return r
                
    # 수식이 작성된 첫 번째 비데이터 행 탐색 (A열이 비어있고 E열이 수식인 경우)
    for r in range(6, 40):
        val_a = sheet.cell(r, 1).value
        val_e = sheet.cell(r, 5).value
        if val_e and isinstance(val_e, str) and val_e.startswith("="):
            if val_a is None or not str(val_a).isdigit():
                return r
                
    return 26 # 기본값

for sheet_name in wb_formulas.sheetnames:
    if "5월 종합" in sheet_name:
        continue
    sheet = wb_formulas[sheet_name]
    sum_row = find_sum_row_formulas(sheet)
    original_sum_rows[sheet_name] = sum_row

wb_formulas.close()

# 2. 원본 파일에서 데이터 추출 (data_only=True로 실제 값 추출)
wb_src = openpyxl.load_workbook(src_path, data_only=True)
migrated_data = {}

for sheet_name in wb_src.sheetnames:
    if "5월 종합" in sheet_name:
        continue
    sheet = wb_src[sheet_name]
    sum_row_src = original_sum_rows.get(sheet_name, 26)
    
    day_data = []
    # 데이터 영역: 6행부터 sum_row_src 이전까지 추출
    for r in range(6, sum_row_src):
        p1_name = sheet.cell(r, 2).value  # B열
        p1_card = sheet.cell(r, 5).value  # E열
        p1_cash = sheet.cell(r, 6).value  # F열
        
        p2_name = sheet.cell(r, 8).value  # H열
        p2_card = sheet.cell(r, 9).value  # I열
        p2_cash = sheet.cell(r, 10).value # J열
        
        margin = sheet.cell(r, 14).value  # N열
        
        day_data.append({
            "p1_name": p1_name, "p1_card": p1_card, "p1_cash": p1_cash,
            "p2_name": p2_name, "p2_card": p2_card, "p2_cash": p2_cash,
            "margin": margin
        })
        
    migrated_data[sheet_name] = day_data

wb_src.close()

# 3. 새 워크북 생성
wb_dest = openpyxl.Workbook()

# 스타일 정의
font_title = Font(name="맑은 고딕", size=18, bold=True, color="000000") # 제목: 검정색 글씨
font_date = Font(name="맑은 고딕", size=10, bold=True)
font_header = Font(name="맑은 고딕", size=10, bold=True, color="000000") # 헤더: 검정색 글씨

# 데이터 셀 폰트
font_data_black = Font(name="맑은 고딕", size=10, color="000000")
font_data_red = Font(name="맑은 고딕", size=10, color="FF0000") # 카드: 빨간색
font_data_blue = Font(name="맑은 고딕", size=10, color="0000FF") # 현금: 파란색

# 합계 셀 폰트 (데이터 셀보다 더 크고 굵게 처리하여 강조)
font_total_black = Font(name="맑은 고딕", size=10, bold=True, color="000000")
font_total_red = Font(name="맑은 고딕", size=10, bold=True, color="FF0000") # 카드 합계: 빨간색 볼드
font_total_blue = Font(name="맑은 고딕", size=10, bold=True, color="0000FF") # 현금 합계: 파란색 볼드

fill_header_dash = PatternFill(start_color="E46C0A", end_color="E46C0A", fill_type="solid") # 대시보드 헤더만 주황색 배경 유지
fill_summary = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid") # 대시보드 합계 배경

align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# 선 스타일
thin_side = Side(style='thin', color='D9D9D9')
medium_side = Side(style='medium', color='E46C0A') # 주황색 테두리
double_side = Side(style='double', color='000000')

def style_range(ws, cell_range, border=None, fill=None, font=None, alignment=None):
    """병합된 셀 범위 전체에 스타일을 적용하는 헬퍼 함수"""
    for row in ws[cell_range]:
        for cell in row:
            if border: cell.border = border
            if fill is not None: cell.fill = fill
            if font: cell.font = font
            if alignment: cell.alignment = alignment

def apply_borders_and_styles(ws, start_row, end_row, start_col, end_col, n1):
    """일일 매출일지의 데이터 및 합계 영역에 테두리 및 폰트 색상을 입히고, 금액은 우측정렬, 글자는 가운데 정렬합니다. (바탕색 없음)"""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            
            # 정렬 제어: 금액(E, F, G, I, J, K, L, N, O열)은 우측정렬, 텍스트(B, H열 및 번호 A열 등)는 가운데정렬
            if c in [5, 6, 7, 9, 10, 11, 12, 14, 15]:
                cell.alignment = align_right
            else:
                cell.alignment = align_center
            
            # 폰트 색상 분리 (카드: 빨강, 현금: 파랑, 나머지: 검정)
            # 카드 열: 5(E), 9(I) / 현금 열: 6(F), 10(J)
            if r >= n1:
                # 합계 및 연산 행 (Row n1 ~ n1+3)
                if c in [5, 9]:
                    cell.font = font_total_red
                elif c in [6, 10]:
                    cell.font = font_total_blue
                else:
                    cell.font = font_total_black
            else:
                # 일반 데이터 행
                if c in [5, 9]:
                    cell.font = font_data_red
                elif c in [6, 10]:
                    cell.font = font_data_blue
                else:
                    cell.font = font_data_black
            
            # 일일 매출일지의 셀 바탕색은 완전히 제거합니다
            cell.fill = PatternFill(fill_type=None)
            
            # 테두리 설정 (외곽라인 깨짐 방지)
            left = thin_side
            right = thin_side
            top = thin_side
            bottom = thin_side
            
            if r == start_row: top = medium_side
            if r == end_row: bottom = medium_side
            if c == start_col: left = medium_side
            if c == end_col: right = medium_side
            
            # 회계 합계선 적용 (Row n1)
            if r == n1:
                top = Side(style='thin', color='000000')
                bottom = double_side
                
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

# 시트별 누적 계산 및 대시보드 링킹을 위한 동적 행 번호 매핑
cumulative_row_map = {}
totals_row_map = {}
daily_sums_row_map = {}

# 4. 31개 일자별 시트 생성 및 데이터 이식
for day in range(1, 32):
    sheet_name = f"5월{day}일"
    ws = wb_dest.create_sheet(sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # 데이터 크기 동적 산출 (최소 20행 보장)
    day_data = migrated_data.get(sheet_name, [])
    num_items = len(day_data)
    num_rows = max(20, num_items)
    
    # A1:O3 병합 대형 타이틀 (바탕색 없음, 검정색 bold 글씨)
    ws.merge_cells("A1:O3")
    ws["A1"] = "(05월)  일    일     매     출     일     지"
    style_range(ws, "A1:O3", fill=PatternFill(fill_type=None), font=font_title, alignment=align_center)
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    
    # 날짜 입력 (K4)
    days_kr = ["월", "화", "수", "목", "금", "토", "일"]
    d = datetime.date(2026, 5, day)
    day_str = days_kr[d.weekday()]
    ws["K4"] = f"2026-05-{day:02d}({day_str})"
    ws["K4"].font = font_date
    ws["K4"].alignment = align_center
    ws.row_dimensions[4].height = 20
    
    # 헤더 입력 및 병합 (Row 5 - 바탕색 없음, 10pt 검정색 bold 글씨)
    ws.merge_cells("B5:D5")
    ws.merge_cells("N5:O5")
    
    headers_daily = {
        1: "No", 2: "품          명", 5: "카  드", 6: "현  금", 7: "합  계",
        8: "품   명", 9: "카  드", 10: "현  금", 11: "합  계", 12: "일 합 계", 14: "마  진"
    }
    
    for col_idx in range(1, 16):
        cell = ws.cell(row=5, column=col_idx)
        cell.font = font_header
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = align_center
        if col_idx in headers_daily:
            cell.value = headers_daily[col_idx]
            
    ws.row_dimensions[5].height = 25
    
    # 데이터 입력 (Row 6 ~ 5+num_rows)
    for i in range(num_rows):
        row = 6 + i
        item = day_data[i] if i < len(day_data) else {}
        
        ws.merge_cells(f"B{row}:D{row}")
        ws.merge_cells(f"N{row}:O{row}")
        
        ws.cell(row=row, column=1, value=i+1)
        ws.cell(row=row, column=2, value=item.get("p1_name"))
        
        # 0 출력을 숨기기 위해 Excel 커스텀 넘버 포맷 적용 (#,##0;-#,##0;;)
        ws.cell(row=row, column=5, value=item.get("p1_card")).number_format = "#,##0;-#,##0;;"
        ws.cell(row=row, column=6, value=item.get("p1_cash")).number_format = "#,##0;-#,##0;;"
        ws.cell(row=row, column=7, value=f"=E{row}+F{row}").number_format = "#,##0;-#,##0;;"
        
        ws.cell(row=row, column=8, value=item.get("p2_name"))
        ws.cell(row=row, column=9, value=item.get("p2_card")).number_format = "#,##0;-#,##0;;"
        ws.cell(row=row, column=10, value=item.get("p2_cash")).number_format = "#,##0;-#,##0;;"
        ws.cell(row=row, column=11, value=f"=I{row}+J{row}").number_format = "#,##0;-#,##0;;"
        
        ws.cell(row=row, column=14, value=item.get("margin")).number_format = "#,##0;-#,##0;;"
        ws.row_dimensions[row].height = 20
        
    # 합계 행 기입 (Row n1)
    n1 = 6 + num_rows
    ws.merge_cells(f"B{n1}:D{n1}")
    ws.cell(row=n1, column=2, value="합   계").alignment = align_center
    ws.cell(row=n1, column=5, value=f"=SUM(E6:E{5+num_rows})").number_format = "#,##0"
    ws.cell(row=n1, column=6, value=f"=SUM(F6:F{5+num_rows})").number_format = "#,##0"
    ws.cell(row=n1, column=7, value=f"=E{n1}+F{n1}").number_format = "#,##0"
    
    ws.cell(row=n1, column=9, value=f"=SUM(I6:I{5+num_rows})").number_format = "#,##0"
    ws.cell(row=n1, column=10, value=f"=SUM(J6:J{5+num_rows})").number_format = "#,##0"
    ws.cell(row=n1, column=11, value=f"=I{n1}+J{n1}").number_format = "#,##0"
    
    ws.cell(row=n1, column=12, value=f"=G{n1}+K{n1}").number_format = "#,##0"
    ws.cell(row=n1, column=14, value=f"=SUM(N6:N{5+num_rows})").number_format = "#,##0"
    ws.row_dimensions[n1].height = 22
    
    # 일별합계 행 기입 (Row n2)
    n2 = 7 + num_rows
    ws.cell(row=n2, column=9, value=f"=E{n1}+I{n1}").number_format = "#,##0"
    ws.cell(row=n2, column=10, value=f"=F{n1}+J{n1}").number_format = "#,##0"
    ws.cell(row=n2, column=12, value=f"=K{n2+1}/{day}").number_format = "#,##0" # 누적합계(n3행) / 경과일수
    ws.row_dimensions[n2].height = 20
    
    # 누적합계 행 기입 (Row n3)
    n3 = 8 + num_rows
    ws.merge_cells(f"L{n3}:M{n3}")
    if day == 1:
        ws.cell(row=n3, column=9, value=f"=I{n2}").number_format = "#,##0"
        ws.cell(row=n3, column=10, value=f"=J{n2}").number_format = "#,##0"
        ws.cell(row=n3, column=11, value=f"=L{n1}").number_format = "#,##0"
        ws.cell(row=n3, column=12, value=f"=L{n1}").number_format = "#,##0"
        ws.cell(row=n3, column=15, value=f"=N{n1}").number_format = "#,##0"
    else:
        prev_sheet = f"5월{day-1}일"
        prev_n3 = cumulative_row_map[day-1]
        ws.cell(row=n3, column=9, value=f"='{prev_sheet}'!I{prev_n3}+I{n2}").number_format = "#,##0"
        ws.cell(row=n3, column=10, value=f"='{prev_sheet}'!J{prev_n3}+J{n2}").number_format = "#,##0"
        ws.cell(row=n3, column=11, value=f"='{prev_sheet}'!K{prev_n3}+L{n1}").number_format = "#,##0"
        ws.cell(row=n3, column=12, value=f"='{prev_sheet}'!L{prev_n3}+L{n1}").number_format = "#,##0"
        ws.cell(row=n3, column=15, value=f"='{prev_sheet}'!O{prev_n3}+N{n1}").number_format = "#,##0"
    ws.row_dimensions[n3].height = 20
    
    # 누적마진율 참고행 (Row n4)
    n4 = 9 + num_rows
    ws.cell(row=n4, column=9, value=f"=I{n3}/K{n3}").number_format = "0.0%"
    ws.row_dimensions[n4].height = 20
    
    # 맵에 동적 행 위치 저장
    cumulative_row_map[day] = n3
    totals_row_map[day] = n1
    daily_sums_row_map[day] = n2
    
    # 테두리 및 폰트(가운데 정렬/우측 정렬 포함) 적용 - 바탕색 없음
    apply_borders_and_styles(ws, 5, n4, 1, 15, n1)
    
    # 열 너비 수동 조정
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 4
    ws.column_dimensions['N'].width = 8
    ws.column_dimensions['O'].width = 8

# 5. 5월 종합 시트 구축
ws_sum = wb_dest.active
ws_sum.title = "5월 종합"
ws_sum.views.sheetView[0].showGridLines = True

ws_sum.merge_cells("A2:D2")
ws_sum["A2"] = "📊 2026년 5월 매출 종합 대시보드"
ws_sum["A2"].font = Font(name="맑은 고딕", size=16, bold=True)
ws_sum.row_dimensions[2].height = 35

# 대시보드는 주황색 헤더와 검정색 폰트 적용
headers = ["날짜", "카드 매출 (KRW)", "현금 매출 (KRW)", "일일 총 합계 (KRW)"]
for col_idx, h in enumerate(headers, 1):
    cell = ws_sum.cell(row=5, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header_dash
    cell.alignment = align_center
    cell.border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=medium_side)
ws_sum.row_dimensions[5].height = 25

for day in range(1, 32):
    row_idx = 5 + day
    sheet_name = f"5월{day}일"
    ws_sum.cell(row=row_idx, column=1, value=f"5월 {day}일").alignment = align_center
    
    n1 = totals_row_map[day]
    n2 = daily_sums_row_map[day]
    
    # 카드 매출: I열의 n2행(일일 카드합계), 현금 매출: J열의 n2행(일일 현금합계), 일일총합계: L열의 n1행(일일총합계)
    ws_sum.cell(row=row_idx, column=2, value=f"='{sheet_name}'!I{n2}").font = font_data_red
    ws_sum.cell(row=row_idx, column=2).number_format = "#,##0"
    ws_sum.cell(row=row_idx, column=2).alignment = align_right # 대시보드 금액 우측정렬
    
    ws_sum.cell(row=row_idx, column=3, value=f"='{sheet_name}'!J{n2}").font = font_data_blue
    ws_sum.cell(row=row_idx, column=3).number_format = "#,##0"
    ws_sum.cell(row=row_idx, column=3).alignment = align_right # 대시보드 금액 우측정렬
    
    ws_sum.cell(row=row_idx, column=4, value=f"='{sheet_name}'!L{n1}").font = font_data_black
    ws_sum.cell(row=row_idx, column=4).number_format = "#,##0"
    ws_sum.cell(row=row_idx, column=4).alignment = align_right # 대시보드 금액 우측정렬
    
    for c in range(1, 5):
        cell = ws_sum.cell(row=row_idx, column=c)
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        # 날짜(1열)는 가운데정렬 유지
        if c == 1:
            cell.alignment = align_center

# 월 종합 합계행 (금액 우측정렬 적용)
tot_row = 37
ws_sum.cell(row=tot_row, column=1, value="월 총 합계").alignment = align_center
ws_sum.cell(row=tot_row, column=2, value="=SUM(B6:B36)").font = font_total_red
ws_sum.cell(row=tot_row, column=2).number_format = "#,##0"
ws_sum.cell(row=tot_row, column=2).alignment = align_right

ws_sum.cell(row=tot_row, column=3, value="=SUM(C6:C36)").font = font_total_blue
ws_sum.cell(row=tot_row, column=3).number_format = "#,##0"
ws_sum.cell(row=tot_row, column=3).alignment = align_right

ws_sum.cell(row=tot_row, column=4, value="=SUM(D6:D36)").font = font_total_black
ws_sum.cell(row=tot_row, column=4).number_format = "#,##0"
ws_sum.cell(row=tot_row, column=4).alignment = align_right

for c in range(1, 5):
    cell = ws_sum.cell(row=tot_row, column=c)
    cell.fill = fill_summary
    cell.border = Border(left=thin_side, right=thin_side, top=Side(style='thin', color='000000'), bottom=double_side)
    if c == 1:
        cell.alignment = align_center

ws_sum.column_dimensions['A'].width = 15
ws_sum.column_dimensions['B'].width = 20
ws_sum.column_dimensions['C'].width = 20
ws_sum.column_dimensions['D'].width = 22

wb_dest.save(dest_path)
print("이전 스타일(금액 우측정렬, 텍스트 가운데정렬, 일일 바탕색 없음)로 원상복구 완료!")
wb_dest.close()
