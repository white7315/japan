import os
import datetime
import openpyxl

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
may_file = os.path.join(desktop, "26년5월 매출일지.xlsx")
june_file = os.path.join(desktop, "26년6월 매출일지.xlsx")

if not os.path.exists(may_file):
    print("Error: 26년5월 매출일지.xlsx not found")
    exit(1)

# 1. 5월 파일 로드
wb = openpyxl.load_workbook(may_file)

# 2. 6월 종합 시트 복사 생성 및 설정
src_dash = wb["5월 종합"]
ws_dash = wb.copy_worksheet(src_dash)
ws_dash.title = "6월 종합"
ws_dash.views.sheetView[0].showGridLines = True

# 대시보드 제목 변경
ws_dash["A2"] = "📊 2026년 6월 매출 종합 대시보드"

# 6월은 30일이므로 May 31st(행 36)를 삭제
ws_dash.delete_rows(36)

# 삭제 후:
# 행 36: 월 총 합계
# 행 37: 카드 / 현금 라벨
# 행 38: 비율 수식

# 일일 날짜 및 시트 연결 수식 쓰기
for day in range(1, 31):
    row_idx = 5 + day
    sheet_name = f"6월{day}일"
    ws_dash.cell(row=row_idx, column=1, value=f"6월 {day}일")
    ws_dash.cell(row=row_idx, column=2, value=f"='{sheet_name}'!I27")
    ws_dash.cell(row=row_idx, column=3, value=f"='{sheet_name}'!J27")
    ws_dash.cell(row=row_idx, column=4, value=f"='{sheet_name}'!L26")
    ws_dash.cell(row=row_idx, column=5, value=f"=COUNTA('{sheet_name}'!B6:B25)+COUNTA('{sheet_name}'!H6:H25)")

# 합계 행 수식 업데이트
ws_dash.cell(row=36, column=2, value="=SUM(B6:B35)")
ws_dash.cell(row=36, column=3, value="=SUM(C6:C35)")
ws_dash.cell(row=36, column=4, value="=SUM(D6:D35)")
ws_dash.cell(row=36, column=5, value="=SUM(E6:E35)")

# 비율 수식 업데이트
ws_dash.cell(row=38, column=2, value="=B36/D36")
ws_dash.cell(row=38, column=3, value="=C36/D36")

# 3. 30개 일자별 시트 생성 및 초기화
template_sheet = wb["5월1일"]
days_kr = ["월", "화", "수", "목", "금", "토", "일"]

for day in range(1, 31):
    sheet_name = f"6월{day}일"
    ws = wb.copy_worksheet(template_sheet)
    ws.title = sheet_name
    ws.views.sheetView[0].showGridLines = True
    
    # 제목 변경
    ws["A1"] = "(06월)  일    일     매     출     일     지"
    
    # 날짜 변경 (K4)
    d = datetime.date(2026, 6, day)
    day_str = days_kr[d.weekday()]
    ws["K4"] = f"2026-06-{day:02d}({day_str})"
    
    # 데이터 입력 영역 비우기 (6행부터 25행까지)
    for row in range(6, 26):
        ws.cell(row=row, column=2).value = None  # 품명1
        ws.cell(row=row, column=5).value = None  # 카드1
        ws.cell(row=row, column=6).value = None  # 현금1
        ws.cell(row=row, column=8).value = None  # 품명2
        ws.cell(row=row, column=9).value = None  # 카드2
        ws.cell(row=row, column=10).value = None # 현금2
        ws.cell(row=row, column=14).value = None # 마진
        
    # 일별합계 평균 분모 업데이트 (L27: K28/day)
    ws.cell(row=27, column=12, value=f"=K28/{day}")
    
    # 누적합계 수식 업데이트 (Row 28)
    if day == 1:
        ws.cell(row=28, column=9, value="=I27")
        ws.cell(row=28, column=10, value="=J27")
        ws.cell(row=28, column=11, value="=L26")
        ws.cell(row=28, column=12, value="=L26")
        ws.cell(row=28, column=15, value="=N26")
    else:
        prev_sheet = f"6월{day-1}일"
        ws.cell(row=28, column=9, value=f"='{prev_sheet}'!I28+I27")
        ws.cell(row=28, column=10, value=f"='{prev_sheet}'!J28+J27")
        ws.cell(row=28, column=11, value=f"='{prev_sheet}'!K28+L26")
        ws.cell(row=28, column=12, value=f"='{prev_sheet}'!L28+L26")
        ws.cell(row=28, column=15, value=f"='{prev_sheet}'!O28+N26")

# 4. 기존 5월 시트 제거
original_sheets = [s for s in wb.sheetnames if "5월" in s or "test_copied" in s]
for s in original_sheets:
    wb.remove(wb[s])

# 6월 종합 시트를 가장 첫 번째 시트로 이동
# openpyxl에서 시트의 순서를 변경하려면 _sheets 리스트를 정렬/재배치합니다.
all_sheets = wb._sheets
june_dash_idx = -1
for idx, s in enumerate(all_sheets):
    if s.title == "6월 종합":
        june_dash_idx = idx
        break

if june_dash_idx != -1:
    june_dash = all_sheets.pop(june_dash_idx)
    all_sheets.insert(0, june_dash)

# 5. 저장
wb.save(june_file)
wb.close()

print("26년6월 매출일지.xlsx 생성 완료!")
