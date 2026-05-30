import os
import datetime
import openpyxl
import shutil
import subprocess

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
workspace = r"E:\japan"

may_file_desktop = os.path.join(desktop, "26년5월 매출일지.xlsx")
june_file_desktop = os.path.join(desktop, "26년6월 매출일지.xlsx")

may_file_ws = os.path.join(workspace, "26년5월 매출일지.xlsx")
june_file_ws = os.path.join(workspace, "26년6월 매출일지.xlsx")

days_kr = ["월", "화", "수", "목", "금", "토", "일"]

# Function to modify dashboard date weekdays and colors
def modify_dashboard(filepath, year, month, total_days):
    print(f"Modifying {os.path.basename(filepath)}...")
    wb = openpyxl.load_workbook(filepath)
    sheetname = f"{month}월 종합"
    if sheetname not in wb.sheetnames:
        print(f"Error: Sheet {sheetname} not found in {filepath}")
        wb.close()
        return False
        
    sheet = wb[sheetname]
    for day in range(1, total_days + 1):
        row_idx = 5 + day
        cell = sheet.cell(row=row_idx, column=1) # Column A
        
        # Calculate date and weekday
        d = datetime.date(year, month, day)
        wd = d.weekday()  # 0:Mon, 5:Sat, 6:Sun
        day_str = days_kr[wd]
        
        # New value: e.g. "5월 1일(금)"
        new_val = f"{month}월 {day}일({day_str})"
        cell.value = new_val
        
        # Determine color (Saturday: Blue, Sunday: Red, Other: Black)
        if wd == 5:
            color_hex = "0000FF"  # Blue
        elif wd == 6:
            color_hex = "FF0000"  # Red
        else:
            color_hex = "000000"  # Black
            
        # Copy font and update color
        curr_font = cell.font
        if curr_font:
            cell.font = openpyxl.styles.Font(
                name=curr_font.name,
                size=curr_font.size,
                bold=curr_font.bold,
                italic=curr_font.italic,
                color=color_hex
            )
        else:
            cell.font = openpyxl.styles.Font(name="맑은 고딕", size=11, color=color_hex)
            
    wb.save(filepath)
    wb.close()
    print(f"Successfully modified {os.path.basename(filepath)}")
    return True

# 1. Modify May Dashboard
if os.path.exists(may_file_desktop):
    modify_dashboard(may_file_desktop, 2026, 5, 31)
else:
    print("Error: May file not found on Desktop")

# 2. Modify June Dashboard
if os.path.exists(june_file_desktop):
    modify_dashboard(june_file_desktop, 2026, 6, 30)
else:
    print("Error: June file not found on Desktop")

# 3. Copy to workspace
print("\nCopying updated files to Git workspace...")
try:
    shutil.copy2(may_file_desktop, may_file_ws)
    print("Copied May journal to workspace.")
    shutil.copy2(june_file_desktop, june_file_ws)
    print("Copied June journal to workspace.")
except Exception as e:
    print(f"Error copying files: {e}")
    exit(1)

# 4. Git commit and push
print("\nSyncing with GitHub...")
git_exe = r"E:\Git\cmd\git.exe"

def run_git(args):
    cmd = [git_exe] + args
    res = subprocess.run(cmd, cwd=workspace, capture_output=True, text=True)
    if res.returncode != 0:
        print(f"Error running git {' '.join(args)}:")
        print("STDERR:", res.stderr)
        return False
    else:
        print(res.stdout)
        return True

run_git(["add", "."])
run_git(["commit", "-m", "Update dashboard dates with colored weekdays (Sat: Blue, Sun: Red)"])
run_git(["push", "-u", "origin", "main"])
print("Git synchronization complete!")
