import os
import openpyxl

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
src_file = "2026년 05월 매출일지.xlsx"
src_path = os.path.join(desktop, src_file)

wb = openpyxl.load_workbook(src_path, data_only=False)

days_to_check = ["5월1일", "5월2일", "5월3일", "5월6일", "5월7일"]

for sheet_name in days_to_check:
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found")
        continue
    sheet = wb[sheet_name]
    print(f"\n--- Formulas for {sheet_name} ---")
    for r in range(26, 30):
        row_vals = []
        for c in range(1, 17):
            val = sheet.cell(r, c).value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                row_vals.append(f"{col_letter}{r}: {val}")
        if row_vals:
            print(f"Row {r}: {', '.join(row_vals)}")

wb.close()
