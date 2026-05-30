import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

desktop = r"C:\Users\dee\OneDrive\바탕 화면"
src_file = "26년5월 매출일지.xlsx"
src_path = os.path.join(desktop, src_file)

wb = openpyxl.load_workbook(src_path)
sheet = wb["5월 종합"]

print("--- Inspecting Styles of Rows 37, 38, 39 in Dashboard ---")
for r in [37, 38, 39]:
    for c in range(1, 5):
        cell = sheet.cell(r, c)
        font = cell.font
        fill = cell.fill
        align = cell.alignment
        border = cell.border
        print(f"\nCell {openpyxl.utils.get_column_letter(c)}{r}: Value='{cell.value}'")
        if font:
            print(f"  Font: size={font.size}, bold={font.bold}, color={font.color.rgb if font.color else None}")
        if fill:
            print(f"  Fill: fill_type={fill.fill_type}, start_color={fill.start_color.rgb if fill.start_color else None}")
        if align:
            print(f"  Alignment: horiz={align.horizontal}, vert={align.vertical}")
        if border:
            print(f"  Border: left={border.left.style if border.left else None}, right={border.right.style if border.right else None}, top={border.top.style if border.top else None}, bottom={border.bottom.style if border.bottom else None}")

wb.close()
